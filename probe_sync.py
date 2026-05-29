"""
probe_sync.py  —  Probe + multi-brand data logger synchronisation and visualiser.

Run with no arguments to open the graphical interface:
    python probe_sync.py

Or pass arguments directly for scripting / automation:
    python probe_sync.py --probe probe.csv --devices a.csv b.csv \\
                         --start "2024-06-01 14:00:00" --output chart.html

Supported file types
---------------------
  .csv  Comma-separated values (most brands)
  .txt  Comma-separated text (same parsing as .csv)
  .xlsx Excel workbook (Omega OM-CP and compatible)

Adding support for a new logger brand
--------------------------------------
  Open logger_formats.json (same folder as this script) and add an entry.
  No changes to this file are required.

Dependencies:
    pip install pandas plotly openpyxl
    tkinter ships with Python (install python3-tk on Linux if needed)
"""

# ── Standard library ──────────────────────────────────────────────────────────
import json
import os
import re
import sys
import threading
from datetime import datetime, timedelta
from pathlib import Path

# ── Third-party ───────────────────────────────────────────────────────────────
import pandas as pd
import plotly.graph_objects as go

# ── Optional GUI ───────────────────────────────────────────────────────────────
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk
    HAS_TK = True
except ImportError:
    HAS_TK = False


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — FORMAT REGISTRY
# ══════════════════════════════════════════════════════════════════════════════

FORMATS_FILE = Path(__file__).parent / "logger_formats.json"

_DEFAULT_FORMATS = {
    "easylog": {
        "description": "Lascar EasyLog / Omega OM-EL-USB",
        "signature":   ["Logger Name:"],
        "date_col":    "Date (DD/MM/YYYY)",
        "time_col":    "Time (HH:mm:ss)",
        "value_col":   "Temperature (\u00b0C)",
        "dayfirst":    True,
    },
    "madgetech": {
        "description": "MadgeTech 4 Data Logger Software",
        "signature":   ["Reading #,Date,Time", "Device Name,Model,Serial Number"],
        "date_col":    "Date",
        "time_col":    "Time",
        "value_col":   "Temperature, \u00b0C",
        "dayfirst":    False,
    },
    "reed": {
        "description": "Reed Instruments data loggers",
        "signature":   ["Device: Reed"],
        "time_col":    "Date/Time",
        "value_col":   "Temperature (\u00b0C)",
        "dayfirst":    False,
    },
    "vaisala": {
        "description": "Vaisala vLog software",
        "signature":   ["Vaisala vLog"],
        "time_col":    "Date/Time",
        "value_col":   "Temperature (\u00b0C)",
        "dayfirst":    False,
    },
    "hobo_onset": {
        "description": "HOBO / Onset USB data logger",
        "signature":   ["Plot Title:"],
        # Column names embed the serial number so we use prefix matching.
        # The actual column will be e.g. "Date Time, GMT-04:00" and
        # "Temp, °C (LGR S/N: ...)" — _find_column() matches by prefix.
        "time_col":    "Date Time",
        "value_col":   "Temp, \u00b0C",
        "datetime_format": "%m/%d/%y %I:%M:%S %p",
    },
    "onset_celsius_txt": {
        "description": "Generic CSV/TXT logger with Celsius column",
        "signature":   ["Celsius("],
        # First column is a row-number with a device-name header — ignored.
        # Time column is "Time", value column starts with "Celsius".
        "time_col":    "Time",
        "value_col":   "Celsius",
    },
    "omega_omcp_xlsx": {
        "description": "Omega OM-CP data logger (Excel export)",
        "signature":   ["OM-CP-"],
        # Date column contains full datetime objects — used directly.
        "time_col":    "Date",
        "value_col":   "Temperature (\u00b0C)",
        "xlsx":        True,
    },
}

def load_registry() -> dict:
    """
    Load format registry from logger_formats.json.
    Creates the file with built-in defaults if it does not exist.
    '_comment' keys are stripped so only real format entries remain.
    """
    if not FORMATS_FILE.exists():
        FORMATS_FILE.write_text(
            json.dumps(_DEFAULT_FORMATS, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(f"Created default format registry: {FORMATS_FILE}")

    with open(FORMATS_FILE, encoding="utf-8") as fh:
        data = json.load(fh)

    return {k: v for k, v in data.items() if not k.startswith("_")}


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — LOW-LEVEL FILE UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

_ENCODINGS = ["utf-8-sig", "utf-8", "latin-1"]   # tried in order


def _detect_encoding(path: str) -> str:
    """
    Return the first encoding in _ENCODINGS that can decode the file's
    first 8 KB without errors.  Falls back to 'latin-1' (which never raises).
    """
    sample = open(path, "rb").read(8192)
    for enc in _ENCODINGS:
        try:
            sample.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "latin-1"


def _read_preamble(path: str, n: int = 20) -> str:
    """Return the first n lines as a single lowercased string."""
    enc = _detect_encoding(path)
    lines = []
    try:
        with open(path, encoding=enc, errors="replace") as fh:
            for _, line in zip(range(n), fh):
                lines.append(line.lower())
    except OSError:
        pass
    return "\n".join(lines)


def _find_header_row(path: str, search_names: list,
                     require_all: bool = False) -> int:
    """
    Scan the file line-by-line and return the index of the header row.

    require_all=False (default, used for auto-detection):
        Return the first line that contains ANY of the search_names as a
        standalone CSV field (surrounded by comma / start / end of line).

    require_all=True (used when loading specific columns):
        Return the first line where ALL search_names appear — either as
        standalone fields OR anywhere as a substring (fallback for quoted
        fields whose names contain internal commas, e.g. HOBO loggers).

    Using AND logic (require_all=True) prevents false matches like the
    Additel 286 preamble row "REF1,,,,," which contains REF1 as a field
    but not "Step Time" — only the real header row contains both.

    Returns 0 if no match is found.
    """
    enc = _detect_encoding(path)

    # Build field-boundary patterns (matches name as a standalone CSV token)
    field_patterns = [
        re.compile(
            r'(?:^|,)\s*"?' + re.escape(s.lower()) + r'"?\s*(?:,|$)',
            re.IGNORECASE,
        )
        for s in search_names if s
    ]

    lines_cache = []
    try:
        with open(path, encoding=enc, errors="replace") as fh:
            for i, line in enumerate(fh):
                stripped = line.strip()
                lines_cache.append(stripped)

                if require_all:
                    # All names must match as fields on this line
                    if all(p.search(stripped) for p in field_patterns):
                        return i
                else:
                    # Any name matches as a field → done
                    if any(p.search(stripped) for p in field_patterns):
                        return i
    except OSError:
        pass

    # ── Fallback: substring matching (handles quoted fields with commas) ──────
    if require_all:
        lower_names = [s.lower() for s in search_names if s]
        for i, line in enumerate(lines_cache):
            if all(name in line.lower() for name in lower_names):
                return i

    return 0


def _find_column(columns, target: str) -> str:
    """
    Locate a column in a DataFrame/Index using three increasingly loose rules:
      1. Exact match
      2. Case-insensitive exact match
      3. Any column whose name starts with `target` (case-insensitive)

    Returns the matched column name, or None if nothing matches.

    This handles HOBO's variable-suffix column names such as
    'Temp, °C (LGR S/N: 22414606, SEN S/N: 22414606)' when the registry
    stores just 'Temp, °C'.
    """
    col_list = list(columns)
    # Exact
    if target in col_list:
        return target
    # Case-insensitive exact
    tl = target.lower()
    for c in col_list:
        if c.lower() == tl:
            return c
    # Starts-with
    for c in col_list:
        if c.lower().startswith(tl):
            return c
    return None


def _read_tabular(path: str, skiprows: int = 0, **kwargs) -> pd.DataFrame:
    """
    Read a delimited data file (.csv or .txt) with automatic encoding and
    separator detection.

    Encoding is detected by _detect_encoding().
    Separator is inferred from the first non-skipped data line:
      comma → default; tab → used if tab count exceeds comma count.
    All extra kwargs are forwarded to pd.read_csv().
    """
    enc = _detect_encoding(path)
    sep = ","
    try:
        with open(path, encoding=enc, errors="replace") as fh:
            for _ in range(skiprows):
                fh.readline()
            sample_line = fh.readline()
        if sample_line.count("\t") > sample_line.count(","):
            sep = "\t"
    except OSError:
        pass
    return pd.read_csv(path, skiprows=skiprows, sep=sep,
                       encoding=enc, **kwargs)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — FORMAT AUTO-DETECTION
# ══════════════════════════════════════════════════════════════════════════════

_TEMP_KEYWORDS    = ["temperature", "temp", "\u00b0c", "\u00b0f",
                      "degc", "degf", "celsius", "fahrenheit"]
_COMBINED_TS_KEYS = ["date/time", "datetime", "timestamp",
                      "date & time", "date-time", "date time"]


def _infer_dayfirst(date_values) -> bool:
    """
    Inspect up to 20 date strings to decide whether the day comes first.
    Returns True (DD/MM) if any first component exceeds 12; False (MM/DD)
    if any second component exceeds 12; False (default) if ambiguous.
    """
    for val in list(date_values)[:20]:
        parts = re.split(r"[/\-.]", str(val).strip().split()[0])
        if len(parts) >= 2:
            try:
                first, second = int(parts[0]), int(parts[1])
                if first  > 12: return True
                if second > 12: return False
            except (ValueError, IndexError):
                continue
    return False


def _read_preamble_xlsx(path: str, n: int = 10) -> str:
    """
    Read the first n rows of an Excel file as a lowercased string so that
    signature matching works the same way for xlsx as for csv/txt.
    openpyxl is used; all cell values are converted to strings.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        lines = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= n:
                break
            lines.append(",".join(str(c) for c in row if c is not None).lower())
        wb.close()
        return "\n".join(lines)
    except Exception:
        return ""


def detect_file_format(path: str, registry: dict) -> tuple:
    """
    Identify which registry entry describes a device file.

    Returns (format_name, format_dict).
    format_name is a registry key or "auto-detected" / "unknown".

    Stage 1 — Signature matching
        The file's first ~20 lines (or xlsx rows) are scanned for each
        entry's signature strings (case-insensitive substring match).

    Stage 2 — Column name matching
        Header columns are compared against registry entries.  Entries
        whose required date_col is absent are disqualified.  Highest
        scoring entry wins.

    Stage 3 — Value pattern analysis
        For unknown brands: timestamp and temperature columns are located
        by keyword scanning; dayfirst is inferred from date values.
    """
    is_xlsx = path.lower().endswith(".xlsx")

    # ── Stage 1: Signature ──────────────────────────────────────────────────
    preamble = _read_preamble_xlsx(path) if is_xlsx else _read_preamble(path)
    for fmt_name, fmt in registry.items():
        for sig in fmt.get("signature", []):
            if sig.lower() in preamble:
                return fmt_name, fmt

    if is_xlsx:
        return "unknown", {}

    # ── Read header for stages 2 & 3 ────────────────────────────────────────
    hints = []
    for fmt in registry.values():
        hints.extend(filter(None, [
            fmt.get("date_col"), fmt.get("time_col"), fmt.get("value_col"),
        ]))
    hints += _COMBINED_TS_KEYS + ["time", "date"]

    header_row = _find_header_row(path, hints, require_all=False)
    try:
        sample = _read_tabular(path, skiprows=header_row, nrows=5, dtype=str)
        sample.columns = sample.columns.str.strip()
        columns = set(sample.columns)
    except Exception:
        return "unknown", {}

    # ── Stage 2: Column name matching ───────────────────────────────────────
    best_score, best_name, best_fmt = 0, None, {}
    for fmt_name, fmt in registry.items():
        date_col  = fmt.get("date_col")
        time_col  = fmt.get("time_col", "")
        value_col = fmt.get("value_col", "")
        score = 0

        if date_col and date_col not in columns:
            continue   # required split-date column absent → disqualify

        if date_col and date_col in columns:  score += 3
        if _find_column(columns, time_col):   score += 2
        if _find_column(columns, value_col):  score += 1

        if score > best_score:
            best_score, best_name, best_fmt = score, fmt_name, fmt

    if best_score > 0:
        return best_name, best_fmt

    # ── Stage 3: Value pattern analysis ─────────────────────────────────────
    inferred = {}
    ts_col  = next((c for c in sample.columns
                    if any(k in c.lower() for k in _COMBINED_TS_KEYS)), None)
    date_col = next((c for c in sample.columns
                     if "date" in c.lower() and "time" not in c.lower()), None)
    time_col = next((c for c in sample.columns
                     if c.lower() == "time" or
                     ("time" in c.lower() and "date" not in c.lower())), None)

    if ts_col:
        inferred["time_col"] = ts_col
    elif date_col and time_col:
        inferred["date_col"] = date_col
        inferred["time_col"] = time_col
        inferred["dayfirst"] = _infer_dayfirst(sample[date_col].dropna())
    elif sample.columns.size > 0:
        inferred["time_col"] = sample.columns[0]

    val_col = next((c for c in sample.columns
                    if any(k in c.lower() for k in _TEMP_KEYWORDS)), None)
    if val_col is None and sample.columns.size > 1:
        val_col = sample.columns[-1]

    inferred["value_col"] = val_col
    return "auto-detected", inferred


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — FILE LOADING
# ══════════════════════════════════════════════════════════════════════════════

# Known probe column names (Additel 286 DAQ)
_PROBE_TIME_CANDIDATES  = ["Step Time", "Time", "Run Time"]
_PROBE_VALUE_CANDIDATES = ["REF1", "REF2",
                            "Temperature (\u00b0C)", "Temperature(\u00b0C)",
                            "Temp (\u00b0C)", "Temp(\u00b0C)", "Temperature"]


def detect_probe_columns(path: str) -> tuple:
    """
    Scan a probe file and return (time_col, available_value_cols).

    Finds the FIRST LINE that contains at least one time-candidate AND at
    least one value-candidate as standalone CSV fields.  This correctly
    skips the Additel 286 metadata rows (e.g. a lone 'REF1,,,,' row) that
    contain only a value candidate but not a time candidate.

    Returns:
      time_col           : matched time column name, or None
      available_value_cols: all REF/temperature columns present in that row
    """
    enc = _detect_encoding(path)

    # Field-boundary patterns for each candidate group
    time_pats = [
        re.compile(r'(?:^|,)\s*"?' + re.escape(s.lower()) + r'"?\s*(?:,|$)',
                   re.IGNORECASE)
        for s in _PROBE_TIME_CANDIDATES
    ]
    value_pats = [
        re.compile(r'(?:^|,)\s*"?' + re.escape(s.lower()) + r'"?\s*(?:,|$)',
                   re.IGNORECASE)
        for s in _PROBE_VALUE_CANDIDATES
    ]

    header_row = 0
    try:
        with open(path, encoding=enc, errors="replace") as fh:
            for i, line in enumerate(fh):
                stripped = line.strip()
                has_time  = any(p.search(stripped) for p in time_pats)
                has_value = any(p.search(stripped) for p in value_pats)
                if has_time and has_value:
                    header_row = i
                    break
    except OSError:
        pass

    try:
        df = _read_tabular(path, skiprows=header_row, nrows=2, dtype=str)
        df.columns = df.columns.str.strip()
    except Exception:
        return None, []

    time_col = next((c for c in _PROBE_TIME_CANDIDATES if c in df.columns), None)
    val_cols = [c for c in _PROBE_VALUE_CANDIDATES if c in df.columns]
    return time_col, val_cols


def _is_absolute_timestamp(value: str) -> bool:
    """
    Return True if a string looks like an absolute datetime rather than an
    elapsed duration.

    Absolute:  "2026-05-27 20:42:00.000"  "2026/05/27 20:42:00"
    Elapsed:   "58:35.0"  "01:03:42"  "3822"

    Rule: if the string contains four or more consecutive digits at the start
    (i.e. a year), it is absolute.  Elapsed values never start with a 4-digit
    group because even a 99-hour elapsed time is only "99:59:59".
    """
    return bool(re.match(r"^\d{4}", value.strip()))


def probe_needs_start_time(path: str) -> bool:
    """
    Return True if the probe file uses elapsed MM:SS format and therefore
    requires a start datetime to anchor its timestamps.

    Returns False for absolute-timestamp files (most common case) and also
    on any read error, so the UI defaults to hiding the field.
    """
    try:
        time_col, _ = detect_probe_columns(path)
        if not time_col:
            return False
        header_row = _find_header_row(path, [time_col], require_all=False)
        df = _read_tabular(path, skiprows=header_row, nrows=2, dtype=str)
        df.columns = df.columns.str.strip()
        if time_col not in df.columns or df.empty:
            return False
        first_val = str(df[time_col].iloc[0]).strip()
        return not _is_absolute_timestamp(first_val)
    except Exception:
        return False


def load_probe(path: str, time_col, value_col,
               start: datetime) -> pd.DataFrame:
    """
    Load the Additel 286 DAQ probe file (.csv or .txt).

    Accepted time columns : "Step Time", "Time", "Run Time"  (auto-detected if None)
    Accepted value columns: "REF1", "REF2", Temperature variants (auto if None)

    The Additel 286 DAQ can export the Step Time column in two formats
    depending on its settings:

    Format A — Elapsed time  (e.g. "58:35.0", resets every hour)
        Parsed to seconds, clock resets stitched, then anchored to `start`.
        The `start` datetime you enter in the GUI is required and used.

    Format B — Absolute datetime  (e.g. "2026-05-27 20:42:00.000")
        Parsed directly as wall-clock timestamps.
        The `start` parameter is still accepted but not needed for alignment
        — the timestamps are already absolute.

    The format is detected automatically by inspecting the first data value.
    """
    def _parse_elapsed(val: str) -> float:
        val = str(val).strip()
        # HH:MM:SS[.d]
        m = re.fullmatch(r"(\d+):(\d{2}):(\d{2}(?:\.\d+)?)", val)
        if m:
            return int(m.group(1))*3600 + int(m.group(2))*60 + float(m.group(3))
        # MM:SS[.d]  (most common Additel 286 export)
        m = re.fullmatch(r"(\d+):(\d{2}(?:\.\d+)?)", val)
        if m:
            return int(m.group(1))*60 + float(m.group(2))
        try:
            return float(val)
        except ValueError:
            raise ValueError(f"Cannot parse elapsed time: '{val}'")

    # ── Auto-detect columns ──────────────────────────────────────────────────
    auto_time, auto_values = detect_probe_columns(path)
    if time_col  is None: time_col  = auto_time
    if value_col is None: value_col = auto_values[0] if auto_values else None

    if not time_col or not value_col:
        auto_t2, auto_v2 = detect_probe_columns(path)
        raise KeyError(
            "Could not identify probe columns automatically.\n"
            f"  Detected time candidates:  {auto_t2}\n"
            f"  Detected value candidates: {auto_v2}\n"
            f"  Expected time:  one of {_PROBE_TIME_CANDIDATES}\n"
            f"  Expected value: one of {_PROBE_VALUE_CANDIDATES}"
        )

    # ── Load ─────────────────────────────────────────────────────────────────
    header_row = _find_header_row(path,
                                  [time_col, value_col],
                                  require_all=True)
    df = _read_tabular(path, skiprows=header_row, dtype=str)
    df.columns = df.columns.str.strip()

    for col in [time_col, value_col]:
        if col not in df.columns:
            raise KeyError(
                f"Probe file: column '{col}' not found.\n"
                f"  Available: {list(df.columns)}"
            )

    # Drop blank / non-data rows
    df = df[df[time_col].notna() & (df[time_col].str.strip() != "")]

    # ── Detect timestamp format from first data value ─────────────────────
    first_val = str(df[time_col].iloc[0]).strip()

    if _is_absolute_timestamp(first_val):
        # Format B — absolute datetimes, parse directly
        df["_abs_time"] = pd.to_datetime(df[time_col].str.strip())
    else:
        # Format A — elapsed MM:SS, stitch resets, anchor to start
        if start is None:
            raise ValueError(
                "START_TIME_REQUIRED: This probe file uses elapsed time format "
                "(e.g. '58:35.0') and needs a start date/time to convert to "
                "real timestamps.\nPlease enter the test start date and time."
            )
        raw_s   = df[time_col].apply(_parse_elapsed)
        offsets = (raw_s.diff() < 0).cumsum() * 3600
        elapsed = raw_s + offsets
        df["_abs_time"] = [start + timedelta(seconds=float(s)) for s in elapsed]

    df["_value"] = pd.to_numeric(df[value_col], errors="coerce")
    return df[["_abs_time", "_value"]].dropna()


def _load_device_xlsx(path: str, fmt: dict) -> pd.DataFrame:
    """
    Load an Excel (.xlsx) device file, e.g. Omega OM-CP exports.

    The Omega OM-CP export has a multi-row metadata preamble followed by a
    header row containing 'Date', 'Time', and 'Temperature (°C)'.  Both the
    Date and Time columns contain full Python datetime objects (openpyxl
    reads them natively), so the Date column is used directly as _abs_time.

    This function is called automatically when the format dict contains
    {"xlsx": true} or when the file extension is .xlsx.
    """
    import openpyxl
    time_col  = fmt.get("time_col", "Date")
    value_col = fmt.get("value_col", "Temperature (\u00b0C)")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find the header row: first row where time_col and value_col both appear
    header_row_idx = None
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        row_strs = [str(c).strip() if c is not None else "" for c in row]
        if time_col in row_strs and value_col in row_strs:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise KeyError(
            f"XLSX file '{os.path.basename(path)}': could not find header row "
            f"containing '{time_col}' and '{value_col}'."
        )

    headers    = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    time_idx   = headers.index(time_col)
    value_idx  = headers.index(value_col)

    records = []
    for row in rows[header_row_idx + 1:]:
        ts  = row[time_idx]
        val = row[value_idx]
        if ts is None or val is None:
            continue
        # openpyxl returns datetime objects directly for date/time cells
        if not isinstance(ts, datetime):
            try:
                ts = pd.to_datetime(str(ts))
            except Exception:
                continue
        try:
            val = float(val)
        except (TypeError, ValueError):
            continue
        records.append({"_abs_time": ts, "_value": val})

    wb.close()
    return pd.DataFrame(records)


def load_device(path: str, fmt: dict) -> pd.DataFrame:
    """
    Load a device logger file using a format dictionary from the registry
    or from auto-detection.

    Dispatches to _load_device_xlsx() for Excel files.
    For CSV/TXT files:
      - Skips the metadata preamble (AND-logic header detection).
      - Uses _find_column() for prefix/startswith matching so that columns
        whose names include variable suffixes (e.g. HOBO serial numbers)
        are found even when the registry stores only the prefix.
      - Builds combined timestamp string for split date+time columns.
      - Parses timestamps; applies dayfirst or explicit format as needed.
    """
    # Excel files have their own loader
    if path.lower().endswith(".xlsx") or fmt.get("xlsx"):
        return _load_device_xlsx(path, fmt)

    time_col        = fmt.get("time_col", "")
    date_col        = fmt.get("date_col")
    value_col       = fmt.get("value_col", "")
    dayfirst        = fmt.get("dayfirst", False)
    datetime_format = fmt.get("datetime_format")

    # Use AND-logic so only the real header row (containing ALL key columns)
    # is returned — prevents false matches in multi-row preambles.
    search = [s for s in [time_col, date_col, value_col] if s]
    header_row = _find_header_row(path, search, require_all=True)

    df = _read_tabular(path, skiprows=header_row, dtype=str)
    df.columns = df.columns.str.strip()

    # Resolve actual column names (handles prefix/startswith matching)
    actual_time  = _find_column(df.columns, time_col)  if time_col  else None
    actual_date  = _find_column(df.columns, date_col)  if date_col  else None
    actual_value = _find_column(df.columns, value_col) if value_col else None

    # Report clear errors if required columns are missing
    for required, label in [(actual_time, time_col), (actual_value, value_col)]:
        if required is None:
            raise KeyError(
                f"'{os.path.basename(path)}': could not find column matching "
                f"'{label}'.\n"
                f"  Available columns: {list(df.columns)}\n"
                f"  Tip: check logger_formats.json for this brand's column names."
            )

    # Build combined timestamp string
    if actual_date:
        combined = df[actual_date].str.strip() + " " + df[actual_time].str.strip()
    else:
        combined = df[actual_time].str.strip()

    # Parse timestamps
    if datetime_format:
        df["_abs_time"] = pd.to_datetime(combined, format=datetime_format)
    else:
        df["_abs_time"] = pd.to_datetime(combined, dayfirst=dayfirst)

    df["_value"] = pd.to_numeric(df[actual_value], errors="coerce")
    return df[["_abs_time", "_value"]].dropna()


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — TIME WINDOW ALIGNMENT
# ══════════════════════════════════════════════════════════════════════════════

def clip_to_probe_window(device_df: pd.DataFrame, name: str,
                         probe_start: datetime, probe_end: datetime,
                         log_fn=print) -> pd.DataFrame:
    """
    Remove device rows outside [probe_start, probe_end] and report trims.
    """
    total = len(device_df)
    mask  = (
        (device_df["_abs_time"] >= probe_start) &
        (device_df["_abs_time"] <= probe_end)
    )
    clipped = device_df[mask].copy()
    dropped  = total - len(clipped)

    if dropped:
        dev_start = device_df["_abs_time"].min()
        dev_end   = device_df["_abs_time"].max()
        early_s   = max(0.0, (probe_start - dev_start).total_seconds())
        late_s    = max(0.0, (dev_end     - probe_end).total_seconds())
        detail    = []
        if early_s: detail.append(f"{early_s:.0f} s before probe start")
        if late_s:  detail.append(f"{late_s:.0f} s after probe end")
        log_fn(f"  \u26a0  '{name}': trimmed {dropped} row(s) "
               f"({', '.join(detail)})")
    else:
        log_fn(f"  \u2713  '{name}': fully within probe window")

    return clipped


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — CHART BUILDING
# ══════════════════════════════════════════════════════════════════════════════

def build_figure(probe_df, device_dfs, device_names):
    """
    Build a Plotly Figure with one trace per dataset.
    Probe → thin line (4-s resolution, no dot markers).
    Devices → thicker line + dot markers (1-min resolution).
    """
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=probe_df["_abs_time"], y=probe_df["_value"],
        mode="lines", name="Probe (ADT286)",
        line=dict(width=1.5, color="#00b4d8"),
    ))

    colors = [
        "#ef233c","#f77f00","#2dc653","#9b5de5",
        "#f15bb5","#fee440","#00bbf9","#fb5607",
        "#e63946","#2a9d8f","#e9c46a","#264653",
    ]
    for i, (df, name) in enumerate(zip(device_dfs, device_names)):
        fig.add_trace(go.Scatter(
            x=df["_abs_time"], y=df["_value"],
            mode="lines+markers", name=name,
            line=dict(width=2, color=colors[i % len(colors)]),
            marker=dict(size=5),
        ))

    fig.update_layout(
        title="Temperature \u2014 Probe + Data Loggers (Synchronized)",
        xaxis_title="Time", yaxis_title="Temperature (\u00b0C)",
        template="plotly_dark", hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="right", x=1),
        font=dict(family="monospace", size=12),
    )
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — CORE PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def run_pipeline(probe_path, probe_time_col, probe_value_col,
                 start_dt, device_paths, output_path,
                 registry, log_fn=print) -> bool:
    """
    Full pipeline: load probe → load & detect devices → align → chart → save.
    Returns True on success, False on error (error is logged via log_fn).
    """
    try:
        log_fn(f"Loading probe: {os.path.basename(probe_path)}")
        probe_df = load_probe(probe_path, probe_time_col,
                              probe_value_col, start_dt)
        log_fn(f"  {len(probe_df):,} rows  |  "
               f"{probe_df['_abs_time'].min():%Y-%m-%d %H:%M:%S} \u2192 "
               f"{probe_df['_abs_time'].max():%Y-%m-%d %H:%M:%S}")

        device_dfs, device_names = [], []
        for path in device_paths:
            name = os.path.splitext(os.path.basename(path))[0]
            fmt_name, fmt = detect_file_format(path, registry)
            desc = registry.get(fmt_name, {}).get("description", fmt_name)
            log_fn(f"Loading: {os.path.basename(path)}  [{desc}]")
            df = load_device(path, fmt)
            log_fn(f"  {len(df):,} rows  |  "
                   f"{df['_abs_time'].min():%Y-%m-%d %H:%M:%S} \u2192 "
                   f"{df['_abs_time'].max():%Y-%m-%d %H:%M:%S}")
            device_dfs.append(df)
            device_names.append(name)

        probe_start = probe_df["_abs_time"].min()
        probe_end   = probe_df["_abs_time"].max()
        log_fn(f"\nAligning to probe window: "
               f"{probe_start:%Y-%m-%d %H:%M:%S} \u2192 "
               f"{probe_end:%Y-%m-%d %H:%M:%S}")

        device_dfs = [
            clip_to_probe_window(df, name, probe_start, probe_end, log_fn)
            for df, name in zip(device_dfs, device_names)
        ]
        for df, name in zip(device_dfs, device_names):
            if df.empty:
                log_fn(f"  \u2717  WARNING: '{name}' has no data in probe window.")

        log_fn("\nBuilding chart\u2026")
        fig = build_figure(probe_df, device_dfs, device_names)
        fig.write_html(output_path)
        log_fn(f"\u2713 Done!  Chart saved to: {output_path}")
        return True

    except Exception as exc:
        import traceback
        log_fn(f"\n\u2717 Error: {exc}")
        log_fn(traceback.format_exc())
        return False


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — GRAPHICAL INTERFACE
# ══════════════════════════════════════════════════════════════════════════════

if HAS_TK:
    class SyncApp(tk.Tk):
        def __init__(self, registry):
            super().__init__()
            self.registry       = registry
            self.device_paths   = []
            self.device_formats = []
            self.title("Probe + Logger Sync Tool")
            self.minsize(660, 600)
            self.resizable(True, True)
            self._build_ui()

        def _build_ui(self):
            PAD = dict(padx=8, pady=4)

            # ── Probe ─────────────────────────────────────────────────────────
            pf = ttk.LabelFrame(self, text=" Probe (Additel 286 DAQ) ")
            pf.pack(fill="x", padx=10, pady=(10, 4))
            pf.columnconfigure(1, weight=1)

            ttk.Label(pf, text="File:").grid(row=0, column=0, sticky="w", **PAD)
            self.probe_var = tk.StringVar()
            ttk.Entry(pf, textvariable=self.probe_var).grid(
                row=0, column=1, sticky="ew", **PAD)
            ttk.Button(pf, text="Browse\u2026",
                       command=self._browse_probe).grid(row=0, column=2, **PAD)

            # Start time row — hidden by default, shown only when the probe
            # file uses elapsed MM:SS format and needs anchoring.
            self._start_label = ttk.Label(pf, text="Start time:")
            self.start_var    = tk.StringVar(value="")
            self._start_entry = ttk.Entry(pf, textvariable=self.start_var,
                                          width=22)
            self._start_hint  = ttk.Label(pf,
                                          text="YYYY-MM-DD HH:MM:SS  "
                                               "(required \u2014 probe uses elapsed time)",
                                          foreground="#e07b00")
            # Don't grid them yet — _show_start_time() does that on demand.

            ttk.Label(pf, text="Channel:").grid(
                row=2, column=0, sticky="w", **PAD)
            self.probe_val_var = tk.StringVar(value="REF1")
            self.channel_combo = ttk.Combobox(
                pf, textvariable=self.probe_val_var,
                values=["REF1", "REF2"], state="readonly", width=12)
            self.channel_combo.grid(row=2, column=1, sticky="w", **PAD)
            ttk.Label(pf, text="(auto-updated when you select a file)",
                      foreground="gray").grid(row=2, column=2, sticky="w", **PAD)

            # ── Devices ───────────────────────────────────────────────────────
            df_ = ttk.LabelFrame(self, text=" Device Loggers ")
            df_.pack(fill="both", expand=True, padx=10, pady=4)

            cols = ("file", "format")
            self.tree = ttk.Treeview(df_, columns=cols,
                                     show="headings", height=7)
            self.tree.heading("file",   text="File")
            self.tree.heading("format", text="Format Detected")
            self.tree.column("file",   width=280, stretch=True)
            self.tree.column("format", width=250, stretch=True)
            sb = ttk.Scrollbar(df_, orient="vertical",
                               command=self.tree.yview)
            self.tree.configure(yscrollcommand=sb.set)
            self.tree.pack(side="left", fill="both",
                           expand=True, padx=(8,0), pady=6)
            sb.pack(side="right", fill="y", pady=6, padx=(0,8))

            bf = ttk.Frame(self)
            bf.pack(fill="x", padx=10, pady=2)
            ttk.Button(bf, text="Add Files\u2026",
                       command=self._add_devices).pack(side="left", padx=4)
            ttk.Button(bf, text="Remove Selected",
                       command=self._remove_selected).pack(side="left", padx=4)
            ttk.Button(bf, text="Clear All",
                       command=self._clear_devices).pack(side="left", padx=4)

            # ── Output ────────────────────────────────────────────────────────
            of = ttk.LabelFrame(self, text=" Output ")
            of.pack(fill="x", padx=10, pady=4)
            of.columnconfigure(1, weight=1)
            ttk.Label(of, text="Save chart to:").grid(
                row=0, column=0, sticky="w", **PAD)
            self.output_var = tk.StringVar(value="chart.html")
            ttk.Entry(of, textvariable=self.output_var).grid(
                row=0, column=1, sticky="ew", **PAD)
            ttk.Button(of, text="Browse\u2026",
                       command=self._browse_output).grid(row=0, column=2, **PAD)

            ttk.Button(self, text="\u25b6  Generate Chart",
                       command=self._run).pack(pady=8)

            lf = ttk.LabelFrame(self, text=" Log ")
            lf.pack(fill="both", expand=True, padx=10, pady=(0,10))
            self.log_box = scrolledtext.ScrolledText(
                lf, height=9, font=("Courier", 9), state="disabled",
                background="#1e1e1e", foreground="#d4d4d4",
                insertbackground="#d4d4d4")
            self.log_box.pack(fill="both", expand=True, padx=4, pady=4)
            self._log("Ready.  Select files and click \u25b6 Generate Chart.")

        def _show_start_time(self):
            """Grid the start time row into the probe frame (row 1)."""
            PAD = dict(padx=8, pady=4)
            self._start_label.grid(row=1, column=0, sticky="w", **PAD)
            self._start_entry.grid(row=1, column=1, sticky="w", **PAD)
            self._start_hint.grid( row=1, column=2, sticky="w", **PAD)

        def _hide_start_time(self):
            """Remove the start time row from the probe frame."""
            self._start_label.grid_remove()
            self._start_entry.grid_remove()
            self._start_hint.grid_remove()

        # ── File dialogs ──────────────────────────────────────────────────────

        def _browse_probe(self):
            path = filedialog.askopenfilename(
                title="Select Additel 286 Probe File",
                filetypes=[("Data files","*.csv *.txt"),
                            ("CSV","*.csv"),("Text","*.txt"),
                            ("All","*.*")],
            )
            if not path:
                return
            self.probe_var.set(path)
            default_out = str(Path(path).parent / "chart.html")
            if self.output_var.get() in ("", "chart.html"):
                self.output_var.set(default_out)
            # Populate channel dropdown from actual file columns
            try:
                _, val_cols = detect_probe_columns(path)
                if val_cols:
                    self.channel_combo["values"] = val_cols
                    self.probe_val_var.set(val_cols[0])
            except Exception:
                pass
            # Show start time field only if the probe uses elapsed MM:SS format
            if probe_needs_start_time(path):
                self._show_start_time()
            else:
                self._hide_start_time()

        def _add_devices(self):
            paths = filedialog.askopenfilenames(
                title="Select Device Logger Files",
                filetypes=[("Data files","*.csv *.txt *.xlsx"),
                            ("CSV","*.csv"),("Text","*.txt"),
                            ("Excel","*.xlsx"),("All","*.*")],
            )
            for path in paths:
                if path in self.device_paths:
                    continue
                fmt_name, fmt = detect_file_format(path, self.registry)
                desc = self.registry.get(fmt_name, {}).get(
                    "description", fmt_name)
                self.device_paths.append(path)
                self.device_formats.append(fmt)
                self.tree.insert("", "end", iid=path,
                                 values=(os.path.basename(path), desc))

        def _remove_selected(self):
            for iid in self.tree.selection():
                self.tree.delete(iid)
                if iid in self.device_paths:
                    idx = self.device_paths.index(iid)
                    self.device_paths.pop(idx)
                    self.device_formats.pop(idx)

        def _clear_devices(self):
            self.tree.delete(*self.tree.get_children())
            self.device_paths.clear()
            self.device_formats.clear()

        def _browse_output(self):
            path = filedialog.asksaveasfilename(
                title="Save Chart As",
                defaultextension=".html",
                filetypes=[("HTML","*.html"),("All","*.*")],
            )
            if path:
                self.output_var.set(path)

        # ── Log helpers ───────────────────────────────────────────────────────

        def _log(self, msg):
            self.after(0, self._log_main, msg)

        def _log_main(self, msg):
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")

        # ── Run ───────────────────────────────────────────────────────────────

        def _run(self):
            probe_path  = self.probe_var.get().strip()
            output_path = self.output_var.get().strip()

            if not probe_path:
                messagebox.showwarning("Missing Input",
                                       "Please select a probe file.")
                return
            if not self.device_paths:
                messagebox.showwarning("Missing Input",
                                       "Please add at least one device file.")
                return

            # Parse start time only if the field is currently visible
            start_dt = None
            if self._start_label.winfo_ismapped():
                start_str = self.start_var.get().strip()
                if not start_str:
                    messagebox.showwarning("Missing Input",
                                           "This probe file needs a start time.\n"
                                           "Please fill in the Start time field.")
                    return
                try:
                    start_dt = datetime.strptime(start_str, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    messagebox.showerror(
                        "Invalid Input",
                        "Start time must be YYYY-MM-DD HH:MM:SS\n"
                        f"Got: {start_str!r}")
                    return

            self.log_box.configure(state="normal")
            self.log_box.delete("1.0", "end")
            self.log_box.configure(state="disabled")

            registry = self.registry

            def worker():
                success = run_pipeline(
                    probe_path      = probe_path,
                    probe_time_col  = None,
                    probe_value_col = self.probe_val_var.get().strip(),
                    start_dt        = start_dt,
                    device_paths    = list(self.device_paths),
                    output_path     = output_path,
                    registry        = registry,
                    log_fn          = self._log,
                )
                if success:
                    self.after(0, lambda: messagebox.showinfo(
                        "Done", f"Chart saved to:\n{output_path}"))
                else:
                    # If the pipeline failed because start time was needed,
                    # reveal the field so the user can fill it in and retry.
                    self.after(0, self._check_if_start_time_needed)

            threading.Thread(target=worker, daemon=True).start()

        def _check_if_start_time_needed(self):
            """Show the start time field if the log contains a start-time error."""
            log_content = self.log_box.get("1.0", "end")
            if "START_TIME_REQUIRED" in log_content:
                self._show_start_time()
                self._log(
                    "\n\u26a0  Start time required.  Fill in the field above "
                    "and click \u25b6 Generate Chart again."
                )


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 9 — CLI FALLBACK
# ══════════════════════════════════════════════════════════════════════════════

def cli_main(registry):
    import argparse
    parser = argparse.ArgumentParser(
        description="Sync Additel 286 probe + device loggers → HTML chart.")
    parser.add_argument("--probe",   required=True)
    parser.add_argument("--devices", required=True, nargs="+")
    parser.add_argument("--start",   required=True,
                        help='"YYYY-MM-DD HH:MM:SS"')
    parser.add_argument("--probe-time-col",  default=None,
                        help="Auto-detected if omitted")
    parser.add_argument("--probe-value-col", default=None,
                        help="Auto-detected (REF1/REF2/Temperature) if omitted")
    parser.add_argument("--output", default="chart.html")
    args = parser.parse_args()

    try:
        start_dt = datetime.strptime(args.start, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        sys.exit(f"ERROR: --start must be YYYY-MM-DD HH:MM:SS, "
                 f"got: {args.start!r}")

    success = run_pipeline(
        probe_path      = args.probe,
        probe_time_col  = args.probe_time_col,
        probe_value_col = args.probe_value_col,
        start_dt        = start_dt,
        device_paths    = args.devices,
        output_path     = args.output,
        registry        = registry,
    )
    sys.exit(0 if success else 1)


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    registry = load_registry()

    if HAS_TK and len(sys.argv) == 1:
        app = SyncApp(registry)
        app.mainloop()
    else:
        if not HAS_TK and len(sys.argv) == 1:
            print("tkinter unavailable.  Pass --help for CLI usage.")
            sys.exit(1)
        cli_main(registry)
